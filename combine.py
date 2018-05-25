
#!Python

import os
import openpyxl
import configuration

os.chdir(configuration.current_dir)

wb = openpyxl.load_workbook('TestReport.xlsx') 
mysheet = wb.get_active_sheet()
x = 4

for filename in os.listdir():
    if filename.startswith(configuration.project_name):
        print(filename)

        wb1 = openpyxl.load_workbook(filename)
        mysheet1 = wb1.get_active_sheet()
        
        # 最大列
        max_column = 15

        # 找到最大行
        for t in range(3,30):
            if mysheet1['B'+str(t)].value == None:
                max_row = t
                break

        # 遍历excel内容
        for i in range(4,max_row):
            for j in range(2,max_column):
               mysheet.cell(row=x,column=j).value = mysheet1.cell(row=i,column=j).value
            x += 1

wb.save('XXX-' + configuration.project_name + '-测试日报.xlsx')







