#!Python

import openpyxl
import os
import configuration

# 切换到当前目录
os.chdir(configuration.current_dir)

# 待统计项目id
projectName = configuration.project_name

# 建立数据库连接
conn = configuration.conn

# 建立游标
cur = conn.cursor()

# 查询bug
def query_bugs(projectId):
    # 查询项目id
    query_project = 'SELECT id FROM zt_module t where t.type=\'bug\' and t.deleted=\'0\' and t.name=\'' + projectName + '\''
    cur.execute(query_project)
    project_id = cur.fetchone()[0]

    # 查询第一级的模块id、name
    module_id1_dict = {}
    query_module_id1 = 'select id,name from zt_module t where t.deleted=\'0\' and t.parent=\'' + str(
        project_id) + '\''
    cur.execute(query_module_id1)
    module_id1_tuple = cur.fetchall()
    for module_id1 in module_id1_tuple:
        module_id1_dict[module_id1[0]] = module_id1[1]

    # 查询第二级的模块id、name
    for module_id1 in module_id1_dict.keys():
        query_module_id2 = 'select id,name from zt_module t where t.deleted=\'0\' and t.parent=\'' + str(
            module_id1) + '\''
        cur.execute(query_module_id2)
        module_id2_tuple = cur.fetchall()

        # 打开模板
        wb = openpyxl.load_workbook('TestReport.xlsx')
        mysheet = wb.get_active_sheet()
        
        # 写入结果
        if module_id2_tuple:
            i = 0
            for module_id2 in module_id2_tuple:
                itemName = module_id1_dict[module_id1] + '_' + module_id2[1]
                bugs_dict = statistics_bugs(module_id2[0])
                
                # 写入结果
                write_result(mysheet,itemName,bugs_dict,i)
                i += 1
            wb.save(projectName + '-' + module_id1_dict[module_id1] + '-测试日报.xlsx')
            
        else:
            itemName = module_id1_dict[module_id1]
            bugs_dict = statistics_bugs(module_id1)

            # 写入结果
            write_result(mysheet,itemName,bugs_dict,0)
            wb.save(projectName + '-' + module_id1_dict[module_id1] + '-测试日报.xlsx')

# 写入结果
def write_result(my_sheet,item_name,bugs_dict,i):
    mysheet = my_sheet
    itemName = item_name
    bugs_dict = bugs_dict
    i = i
    mysheet.cell(row=4+i,column=2).value = itemName
    for item_name in bugs_dict.keys():
        # 问题总数
        if item_name == 'total_bugs':
            mysheet.cell(row=4 + i, column=3).value = bugs_dict[item_name]
        # 未关闭问题数
        if item_name == 'open_bugs':
            mysheet.cell(row=4 + i, column=4).value = bugs_dict[item_name]
        # 问题分类
        if item_name == 'type_bugs':
            for type_name in bugs_dict['type_bugs'].keys():
                if type_name == 'function ':  # 功能
                    mysheet.cell(
                        row=4 + i,
                        column=7).value = bugs_dict['type_bugs'][type_name]
                if type_name == 'UserExperience':  # 用户体验
                    mysheet.cell(
                        row=4 + i,
                        column=5).value = bugs_dict['type_bugs'][type_name]
                if type_name == 'load':  # 性能
                    mysheet.cell(
                        row=4 + i,
                        column=6).value = bugs_dict['type_bugs'][type_name]
                if type_name == 'compatibility':  # 兼容
                    mysheet.cell(
                        row=4 + i,
                        column=8).value = bugs_dict['type_bugs'][type_name]
                if type_name == 'UI':  # UI
                    mysheet.cell(
                        row=4 + i,
                        column=9).value = bugs_dict['type_bugs'][type_name]
        # 问题严重级别
        if item_name == 'severity_bugs':
            for severity_name in bugs_dict['severity_bugs'].keys():
                if severity_name == 1:  # 致命
                    mysheet.cell(
                        row=4 + i, column=14).value = bugs_dict[
                            'severity_bugs'][severity_name]
                if severity_name == 2:  # 严重
                    mysheet.cell(
                        row=4 + i, column=13).value = bugs_dict[
                            'severity_bugs'][severity_name]
                if severity_name == 3:  # 一般
                    mysheet.cell(
                        row=4 + i, column=12).value = bugs_dict[
                            'severity_bugs'][severity_name]
                if severity_name == 4:  # 提示
                    mysheet.cell(
                        row=4 + i, column=11).value = bugs_dict[
                            'severity_bugs'][severity_name]
                if severity_name == 5:  # 建议
                    mysheet.cell(
                        row=4 + i, column=10).value = bugs_dict[
                            'severity_bugs'][severity_name]


# 统计bug
def statistics_bugs(moduleId):
    #
    bugs_dict = {}

    # 问题总数
    query_total_bugs = 'select count(*) from zt_bug t where t.deleted=\'0\' and t.module =\'' + str(
        moduleId) + '\''
    cur.execute(query_total_bugs)
    total_bugs = cur.fetchone()[0]
    bugs_dict['total_bugs'] = total_bugs

    # 未关闭数
    query_open_bugs = 'select count(*) from zt_bug t where t.deleted=\'0\' and t.status <> \'closed\' and t.module =\'' + str(
        moduleId) + '\''
    cur.execute(query_open_bugs)
    open_bugs = cur.fetchone()[0]
    bugs_dict['open_bugs'] = open_bugs

    # 按问题分类
    type_bugs_dict = {}
    query_type_bugs = 'select type,count(id) from zt_bug t where t.deleted=\'0\' and t.status <> \'closed\' and t.module =\'' + str(
        moduleId) + '\'' + 'group by type'
    cur.execute(query_type_bugs)
    type_bugs_tuple = cur.fetchall()
    for type_bug in type_bugs_tuple:
        type_bugs_dict[type_bug[0]] = type_bug[1]

    bugs_dict['type_bugs'] = type_bugs_dict

    # 按严重程度
    severity_bugs_dict = {}
    query_severity_bugs = 'select severity,count(id) from zt_bug t where t.deleted=\'0\' and t.status <> \'closed\' and t.module =\'' + str(
        moduleId) + '\'' + 'group by severity'
    cur.execute(query_severity_bugs)
    type_bugs_tuple = cur.fetchall()
    for severity_bug in type_bugs_tuple:
        severity_bugs_dict[severity_bug[0]] = severity_bug[1]

    bugs_dict['severity_bugs'] = severity_bugs_dict

    return bugs_dict

# 查询bugs
query_bugs(projectName)

# 释放游标、断开连接
cur.close()
conn.close()